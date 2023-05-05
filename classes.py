import openpyxl
from openpyxl.styles import Alignment, PatternFill


class ExcelFile:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None

    def __enter__(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        return self.workbook

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.workbook:
            self.workbook.save(self.file_path)

    def save(self):
        if self.workbook:
            self.workbook.save(self.file_path)


def align(sheet_cell):
    sheet_cell.alignment = Alignment(horizontal='center', vertical='center')


def color_red(sheet_cell):
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # red fill color
    sheet_cell.fill = red_fill


def color_green(sheet_cell):
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # green fill color
    sheet_cell.fill = green_fill
