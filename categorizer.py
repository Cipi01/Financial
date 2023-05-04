import openpyxl
from openpyxl.styles import Alignment, PatternFill
import string
from categories import taxes, extractions, fuel, deliveries, supermarkets, salary, datorii_in, datorii_out, diverse, \
    fastfood, clothes


def align(sheet_cell):
    sheet_cell.alignment = Alignment(horizontal='center', vertical='center')


workbook = openpyxl.load_workbook('C:/Users/cipri/OneDrive/Desktop/Financial/04_23.xlsx')
source_sheet = workbook['Sheet1']

if 'Sheet2' not in workbook.sheetnames:
    workbook.create_sheet('Sheet2')
sheet2 = workbook['Sheet2']

keyword_categories = {
    'Supermarkets': ['CARREFOUR', 'MEDIAGALAXY', 'PRO SPERANTA', 'LIDL', 'PROFI', 'BAMT', 'Surii Mari', 'MTV',
                     'NADIDA', 'AUCHAN', 'ALTEX', 'KAUFLAND', 'ATTRIUS', 'ILUISI', 'EUR COMTUR'],
    'Salary': ["Alim conventie"],
    'Deliveries': ['EMAG'],
    'Plata datorii': ['Incasare intrab'],
    'Dat datorii': ['Plata i'],
    'Fuel': ['ARAL', 'OMV', 'PETROIL'],
    'Extrageri': ['Retrageri'],
    'Taxes': ['Comision'],
    'Fast-food': ['SELECT', 'ISTANBUL'],
    'Diverse': ['BOLTEU', 'Google Payment'],
    'Clothes': ['NEW YORKER']
}

letters_uppercase = string.ascii_uppercase
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # red fill color
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # green fill color
# Iterate over the keys of the keyword_categories dictionary
for i, key in enumerate(keyword_categories.keys()):
    first_let = letters_uppercase[i * 2]
    last_let = letters_uppercase[i * 2 + 1] if i * 2 + 1 < len(letters_uppercase) else first_let
    sheet2.merge_cells(f"{first_let}1:{last_let}1")
    sheet2[f'{first_let}1'] = f'{key}'
    sheet2[f'{first_let}2'] = 'nume'
    sheet2[f'{last_let}2'] = 'suma'
    varA = sheet2[f'{first_let}1']
    align(varA)
    varB1 = sheet2[f'{first_let}2']
    varB2 = sheet2[f'{last_let}2']
    align(varB1)
    align(varB2)

    if key == "Salary" or key == "Plata datorii":
        varA.fill = green_fill
    else:
        varA.fill = red_fill

supermarkets(keyword_categories, source_sheet, sheet2)
salary(keyword_categories, source_sheet, sheet2)
deliveries(keyword_categories, source_sheet, sheet2)
datorii_in(keyword_categories, source_sheet, sheet2)
datorii_out(keyword_categories, source_sheet, sheet2)
fuel(keyword_categories, source_sheet, sheet2)
extractions(keyword_categories, source_sheet, sheet2)
taxes(keyword_categories, source_sheet, sheet2)
fastfood(keyword_categories, source_sheet, sheet2)
diverse(keyword_categories, source_sheet, sheet2)
clothes(keyword_categories, source_sheet, sheet2)

# Sheet 3
if 'Sheet3' not in workbook.sheetnames:
    workbook.create_sheet('Sheet3')
sheet3 = workbook['Sheet3']

sheet3['A1'] = 'Centralizator'
letters = []
for i in range(len(keyword_categories.keys())):
    letter = letters_uppercase[i]
    letters.append(letter)
sheet3.merge_cells(f'A1:{letters[-1]}1')
align(sheet3['A1'])
for i, key in enumerate(keyword_categories.keys()):
    first_let = letters_uppercase[i]
    second_let = letters_uppercase[i * 2 + 1]
    sheet3[f'{first_let}2'] = f'{key}'
    sheet3[f'{first_let}3'] = f'=SUM(Sheet2!{second_let}3:{second_let}100)'
    varA = sheet3[f'{first_let}2']

    if key == "Salary" or key == "Plata datorii":
        varA.fill = green_fill
    else:
        varA.fill = red_fill

sheet3['A6'] = 'TOTAL'
sheet3.merge_cells('A6:B6')
align(sheet3['A6'])

sheet3['A7'] = 'CREDIT'
sheet3['B7'] = 'DEBIT'
align(sheet3['A7'])
align(sheet3['B7'])

sheet3['A8'] = '=SUMIF(Sheet1!D2:D1000,"Credit",Sheet1!A2:A1000)'
sheet3['B8'] = '=SUMIF(Sheet1!D2:D1000,"Debit",Sheet1!A2:A1000)'
align(sheet3['A8'])
align(sheet3['B8'])
sheet3['A8'].fill = green_fill
sheet3['B8'].fill = red_fill
workbook.save('example.xlsx')
